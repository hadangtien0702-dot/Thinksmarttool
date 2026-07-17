# -*- coding: utf-8 -*-
"""Sinh Product Hub cho Thinksmart Tool: 1 file XLSX (4 tab) + 1 bản Markdown.

Chạy:  python product/build-product-hub.py   (từ repo root)
Output: product/Thinksmart-Product-Hub.xlsx + product/PRODUCT-HUB.md

Quy trình cập nhật: sửa DATA bên dưới → chạy lại script → upload file xlsx mới
lên Google Drive (đè bản cũ) → commit. Nguồn sự thật là FILE NÀY (versioned theo git).
Cập nhật lần cuối: 2026-07-17.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

BRAND = "4F00CA"
HEADER_FILL = PatternFill("solid", fgColor=BRAND)
GROUP_FILL = PatternFill("solid", fgColor="ECE6FE")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
GROUP_FONT = Font(bold=True, size=11, color="3A0E7A")
WRAP = Alignment(wrap_text=True, vertical="top")

VISION = [
    ("Tầm nhìn", "Bộ công cụ trong túi của sale bảo hiểm Thinksmart — tạo tài liệu bán hàng đẹp, "
                 "đúng chuẩn, trong vài phút, ngay cả khi đang ngồi với khách."),
    ("North star", "Thời gian từ 'mở tool' đến 'bản chào nằm trong Zalo của khách' — mục tiêu dưới 5 phút."),
    ("Người dùng", "~100 sale bảo hiểm nhân thọ người Việt tại Mỹ (Thinksmart Insurance, Houston TX). "
                   "Khách của họ: cộng đồng người Việt, tư vấn tại nhà/quán, giao tiếp qua Zalo/Messenger."),
]

PRINCIPLES = [
    ("1. Nhanh gọn", "Mọi tính năng đo bằng số chạm. Sale bị khách gọi cắt ngang liên tục — flow phải sống sót "
                     "qua gián đoạn (dirty-state, tự lưu trước khi xuất, nút theo ngữ cảnh)."),
    ("2. Data khách nằm trên máy sale", "Không server lưu data khách, không bắt đăng nhập. Nháp = localStorage "
                     "trình duyệt từng người. Đây là lợi thế privacy + compliance, là nguyên tắc thiết kế chứ không phải tình cờ."),
    ("3. Một màn hình một việc chính", "Nút đổi theo ngữ cảnh (mẫu gốc → 'Tạo bản cho khách'; bản nháp → 'Lưu Nháp'). "
                     "Không bày mọi nút cùng lúc. Chủ ghét nút thừa."),
    ("4. Trình bày, không phải illustration engine", "Tool KHÔNG BAO GIỜ tự tính phí/cash value — số liệu do sale chép "
                     "từ illustration chính thức của hãng (AG 49-A/B). Disclaimer là mặc định trên mọi bản xuất."),
    ("5. Học tích luỹ mỗi ngày", "Mỗi phiên làm việc để lại changelog + bài học (design-lessons). Mỗi lần push: "
                     "bump version badge + cập nhật skill. Kiến thức compound cho đường dài."),
    ("6. Nền tảng mở rộng dần", "Module mới tái dùng hạ tầng cũ: engine SVG fill-text dùng cho proposal → name card → "
                     "(tương lai) mẫu bài đăng FB; khung thư viện Brochure → kho video training."),
]

# (Nhóm, Tính năng, Giá trị cho sale, Effort, Rủi ro/Điều kiện, Trạng thái)
ROADMAP = [
    ("NOW — làm ngay (mỗi cái vài giờ)", [
        ("Nút 'Gửi cho khách' 1 chạm (Web Share API)", "Gửi JPEG/PDF thẳng vào Zalo/Messenger ngay lúc còn ngồi với khách — khoảnh khắc vàng chốt deal", "S", "Cần fallback tải file cho desktop", "Chưa làm"),
        ("Chế độ trình bày (ẩn UI, full màn hình)", "Đưa điện thoại cho khách xem không lộ nút sửa/nháp — chuyên nghiệp", "S", "Giữ pinch-zoom hoạt động", "Chưa làm"),
        ("Sao lưu / khôi phục nháp ra file JSON", "Chống mất trắng 10 bản nháp khi xoá cache/đổi máy — rủi ro data lớn nhất hiện nay", "S", "Nhắc backup định kỳ kín đáo", "Chưa làm"),
        ("Disclaimer + footer tự động mọi bản xuất", "'Chỉ mang tính minh hoạ — not a carrier illustration' + tên agent, không tắt được. Lá chắn compliance cho cả đội", "S", "Đối chiếu yêu cầu từng hãng trước khi khoá nội dung", "Chưa làm"),
        ("QR code vCard trên name card & proposal", "Khách quét là lưu contact; bản chào chuyển tiếp thành name card lan truyền; kèm số license để khách kiểm chứng", "S", "Không nhúng data khách vào QR", "Chưa làm"),
        ("Thư viện tin nhắn mẫu tiếng Việt (copy 1 chạm)", "Cảm ơn sau gặp / nhắc hẹn / sinh nhật / xin referral — việc 3 phút còn 10 giây, lời lẽ cả đội nhất quán", "S", "Cho sửa 1-2 chữ trước khi copy, tránh 'công nghiệp'", "Chưa làm"),
    ]),
    ("NEXT — kế tiếp (vài ngày mỗi cái)", [
        ("So sánh 2–3 phương án cạnh nhau", "Trả lời 'đóng ít hơn thì sao?' — ghép các nháp thành 1 ảnh so sánh, tâm lý 3 lựa chọn giúp chốt gói giữa", "M", "Số liệu PHẢI chép từ illustration hãng + disclaimer; render 8.8MB x3 cần tuần tự", "Chưa làm"),
        ("Máy tính nhu cầu bảo vệ DIME", "Nợ + thu nhập×năm + nhà + học phí con → mệnh giá có căn cứ, bấm là điền vào ô", "S", "Ghi rõ 'tham khảo, không phải tư vấn tài chính'", "Chưa làm"),
        ("Checklist giấy tờ làm hồ sơ theo hãng", "Gửi khách chuẩn bị trước — hẹn ký 1 lần là xong, deal không nguội", "S", "Nội dung để file JSON dễ sửa khi hãng đổi yêu cầu", "Chưa làm"),
        ("Ghi chú 1 dòng + trạng thái cho từng nháp", "Phân vân / Hẹn lại / Đã chốt — bị cắt ngang 2 ngày vẫn nhớ khách đang ở bước nào", "S", "Giữ đúng 1 note + 1 trạng thái, KHÔNG phình thành CRM", "Chưa làm"),
        ("Bảng tin đội (news.json) + link cài đặt 1 chạm + tour 60s", "Cả đội biết mẫu mới/phí mới ngay khi mở tool; onboarding sale mới bằng 1 link Zalo", "S", "Đưa việc cập nhật news.json vào push-checklist", "Chưa làm"),
        ("Thẻ giải thích thuật ngữ tiếng Việt (IUL, cash value...)", "Khúc khó nhất khi tư vấn cộng đồng — mở thẻ hình đẹp chỉ cho khách thay vì nói suông", "M", "Nội dung trung tính, chủ agency duyệt trước", "Chưa làm"),
        ("Cảnh báo QC mềm trước khi xuất + nhãn Bảo đảm/Không bảo đảm", "Chặn 80% bản chào lỗi (placeholder chưa sửa, thiếu số) + tách rõ số cam kết vs minh hoạ (NAIC #582)", "M", "Chỉ cảnh báo vàng, KHÔNG chặn cứng — sale đang ngồi với khách", "Chưa làm"),
    ]),
    ("LATER — nền tảng (đúng tầm nhìn platform)", [
        ("Mẫu bài đăng Facebook theo hãng", "Nguồn khách số 1; tái dùng engine SVG sẵn có; tên+SĐT tự điền từ preset", "M", "Disclaimer trong mẫu; duyệt bộ đầu bằng skill viet-insurance-ads", "Chưa làm"),
        ("Kho video & tài liệu training", "Onboarding sale mới ngay trong tool; tái dùng khung Brochure + link YouTube unlisted", "M", "Không commit MP4 vào repo; nội dung là việc của chủ", "Chưa làm"),
        ("Field-map theo manifest (hạ tầng)", "Thêm hãng mới = 15 phút gán nhãn thay vì 1-2 ngày dò toạ độ. LÀM TRƯỚC khi thêm Allianz/mẫu mới", "M", "Giữ fallback heuristic cho 4 mẫu cũ, không thì nháp đang dùng vỡ", "Chưa làm"),
        ("Mẫu Whole Life / Final Expense", "Cộng đồng lớn tuổi hỏi nhiều; layout gần IUL nên tái dùng editor", "L", "Làm SAU field-map manifest", "Chưa làm"),
        ("Xuất song ngữ Việt–Anh", "Thuyết phục cả 2 thế hệ trong nhà (con đọc English review giúp bố mẹ)", "M", "Cần thiết kế lại spacing từng mẫu", "Chưa làm"),
        ("Đếm sử dụng siêu nhẹ (Vercel function → Google Sheet)", "Biết đội có dùng không, ai 2 tuần im ắng cần kèm. Chỗ DUY NHẤT đáng làm backend nhẹ", "M", "TUYỆT ĐỐI không log data khách; fail-silent", "Chưa làm"),
    ]),
    ("LẰN RANH — không làm", [
        ("Tự tính phí / cash value dự phóng", "—", "—", "KHÔNG BAO GIỜ: illustration IUL phải từ phần mềm hãng phê duyệt (AG 49-A/B). Tool chỉ trình bày số chép sang", "Nguyên tắc"),
        ("Bảng giá tham khảo tự tra", "—", "—", "Chỉ làm nếu có người CAM KẾT cập nhật số; giá sai mất uy tín + rủi ro compliance", "Hoãn"),
        ("CRM đầy đủ / server chứa data khách", "—", "—", "Không, trừ khi đánh giá lại có chủ đích khi làm module quản lý khách", "Nguyên tắc"),
    ]),
]

# (Version, Ngày, Tính năng chính, Giá trị / ý nghĩa)
RELEASES = [
    ("v1.02", "17/07/2026",
     "Nháp trình duyệt cho site live (draftsMode browser, max 10 bản) · modal dialog đẹp thay alert hệ thống (13+1 chỗ) · tên khách tự điền khi 'Tạo bản cho khách' · fix bug round-trip '$999.99.70'",
     "100 sale dùng đồng thời không cần đăng nhập; data khách an toàn trên máy từng người; đóng PENDING lâu nhất (lưu nháp trên live); bớt 1 bước nhập tay mỗi bản chào."),
    ("v1.01", "17/07/2026",
     "Placeholder chuẩn cho cả 5 mẫu gốc (Nguyen Van An / Texas / (000) 000-0000) · mobile: nút 44px, safe-area iPhone, chặn pull-to-refresh, tap feedback · contrast đạt AA cả 2 theme · pre-push checklist thành quy định",
     "Mẫu gốc chuyên nghiệp, hết lộ SĐT thật của team và data test; dùng mượt trên điện thoại — nơi sale làm việc thật; quy trình release có kỷ luật."),
    ("v1.00", "17/07/2026",
     "Badge version trong UI · 11 font THẬT (fix 'font bị đổi khi xuất' — bộ woff cũ là đồ giả) · custom domain tool.thinksmartinsurance.com",
     "Bản xuất đúng font trên MỌI máy (kể cả máy sale không cài SF Pro); địa chỉ dễ nhớ để chia sẻ cho đội; phân biệt được bản local vs live."),
    ("trước v1.00", "13–16/07/2026",
     "Rebrand Thinksmart Tool · design system tím + light/dark · 3 công cụ (Proposal / Brochure / Name Card) · workflow 4 bước · master protection + clone cho khách · agent preset tự nhớ · mobile drawer/bottom-sheet · module hoá 5 file JS · audit UI/UX toàn tool",
     "Từ editor SVG thô thành sản phẩm nội bộ hoàn chỉnh mà sale thật dùng hằng ngày."),
]

# (Ngày, Bài học)
LESSONS = [
    ("17/07", "Ship xong phải test ROUND-TRIP (lưu → mở lại → so từng field): bug '$999.99 thành $999.99.70' chỉ lộ khi mở lại, nhìn lúc lưu thì mọi thứ đều đẹp."),
    ("17/07", "Snapshot key→value của SVG nhiều tspan phải lưu CẢ DÒNG, và mọi write phải xoá tspan em — nửa vời là data khách sai số tiền."),
    ("17/07", "100 người dùng đồng thời KHÔNG có nghĩa cần backend. Hỏi 'ai giữ state?': nháp cá nhân thuộc máy sale (localStorage), server chỉ phục vụ mẫu đọc chung."),
    ("17/07", "Dialog tự vẽ phải mở ĐỒNG BỘ (force reflow), đừng qua requestAnimationFrame — tab nền bị throttle là dialog kẹt vô hình."),
    ("17/07", "`viewport-fit=cover` mà thiếu env(safe-area-inset-*) là làm nửa việc — nút bấm nằm dưới thanh home iPhone."),
    ("17/07", "Audit touch target phải quét MỌI nút trong vùng, kể cả nút có chữ bị giấu font-size:0 — media query chỉ nâng icon-btn là sót."),
    ("17/07", "Lỗi phát hiện qua accessibility tree phải verify lại bằng textContent thật trước khi sửa — a11y tree có thể hiện attribute title thay vì chữ nhìn thấy."),
    ("16/07", "'Chuẩn hoá' một tool đang dùng hằng ngày = siết độ chính xác, KHÔNG đổi look — bản sắc (tím #4F00CA, chấm bi, 4 bước) đã tiêu đúng chỗ."),
    ("15/07", "Sale hay bị khách gọi cắt ngang → dirty-state + chấm cam + confirm rời trang + export tự lưu. Thiết kế cho người bị gián đoạn, không phải người ngồi yên."),
    ("15/07", "Người dùng thật ghét nút thừa ('4 nút lằng nhằng') → nút theo ngữ cảnh + tự động hoá preset thay vì thêm lựa chọn."),
]


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_xlsx(path):
    wb = Workbook()

    ws = wb.active
    ws.title = "1. Vision & Nguyên tắc"
    ws.append(["Mục", "Nội dung"])
    style_header(ws, 2)
    for k, v in VISION:
        ws.append([k, v])
    ws.append(["", ""])
    ws.append(["NGUYÊN TẮC SẢN PHẨM", ""])
    ws.cell(row=ws.max_row, column=1).font = GROUP_FONT
    ws.cell(row=ws.max_row, column=1).fill = GROUP_FILL
    ws.cell(row=ws.max_row, column=2).fill = GROUP_FILL
    for k, v in PRINCIPLES:
        ws.append([k, v])
    set_widths(ws, [34, 100])
    for r in ws.iter_rows(min_row=2):
        for cell in r:
            cell.alignment = WRAP

    ws = wb.create_sheet("2. Roadmap")
    ws.append(["Nhóm", "Tính năng", "Giá trị cho sale", "Effort", "Rủi ro / Điều kiện", "Trạng thái"])
    style_header(ws, 6)
    for group, items in ROADMAP:
        ws.append([group, "", "", "", "", ""])
        for cell in ws[ws.max_row]:
            cell.fill = GROUP_FILL
            cell.font = GROUP_FONT
        for feat, value, effort, risk, status in items:
            ws.append(["", feat, value, effort, risk, status])
    set_widths(ws, [26, 44, 56, 8, 52, 12])
    for r in ws.iter_rows(min_row=2):
        for cell in r:
            cell.alignment = WRAP

    ws = wb.create_sheet("3. Release Notes")
    ws.append(["Version", "Ngày", "Tính năng chính", "Giá trị / Ý nghĩa"])
    style_header(ws, 4)
    for row in RELEASES:
        ws.append(list(row))
    set_widths(ws, [12, 14, 62, 62])
    for r in ws.iter_rows(min_row=2):
        for cell in r:
            cell.alignment = WRAP

    ws = wb.create_sheet("4. Bài học")
    ws.append(["Ngày", "Bài học"])
    style_header(ws, 2)
    for row in LESSONS:
        ws.append(list(row))
    set_widths(ws, [10, 120])
    for r in ws.iter_rows(min_row=2):
        for cell in r:
            cell.alignment = WRAP

    wb.save(path)


def build_md(path):
    L = []
    L.append("# Thinksmart Tool — Product Hub")
    L.append("")
    L.append("> Bản Markdown đồng bộ với Google Sheet (bản nhìn-cho-người). Nguồn sự thật: `product/build-product-hub.py` — sửa DATA trong đó rồi chạy lại để tái sinh cả hai. Cập nhật: 17/07/2026.")
    L.append(">")
    L.append("> 📊 **Google Sheet:** https://docs.google.com/spreadsheets/d/1Y8kpimASEXucj8a5Mio2BNVabfwaX2ERZBKJ7cPUClQ/edit")
    L.append(">")
    L.append("> 📁 **Drive folder:** https://drive.google.com/drive/folders/1Ngep_IP1gAxkSwZ34daDUc5dGtTNdf6B (tài khoản xuanthuongqtkd@gmail.com)")
    L.append("")
    L.append("## 1. Vision & Nguyên tắc")
    L.append("")
    for k, v in VISION:
        L.append(f"**{k}:** {v}")
        L.append("")
    L.append("### Nguyên tắc sản phẩm")
    L.append("")
    for k, v in PRINCIPLES:
        L.append(f"- **{k}** — {v}")
    L.append("")
    L.append("## 2. Roadmap")
    L.append("")
    for group, items in ROADMAP:
        L.append(f"### {group}")
        L.append("")
        L.append("| Tính năng | Giá trị cho sale | Effort | Rủi ro / Điều kiện | Trạng thái |")
        L.append("|---|---|---|---|---|")
        for feat, value, effort, risk, status in items:
            L.append(f"| {feat} | {value} | {effort} | {risk} | {status} |")
        L.append("")
    L.append("## 3. Release Notes")
    L.append("")
    L.append("| Version | Ngày | Tính năng chính | Giá trị / Ý nghĩa |")
    L.append("|---|---|---|---|")
    for v, d, f, m in RELEASES:
        L.append(f"| {v} | {d} | {f} | {m} |")
    L.append("")
    L.append("## 4. Bài học")
    L.append("")
    for d, lesson in LESSONS:
        L.append(f"- **{d}:** {lesson}")
    L.append("")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(L))


if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    xlsx = os.path.join(here, "Thinksmart-Product-Hub.xlsx")
    md = os.path.join(here, "PRODUCT-HUB.md")
    build_xlsx(xlsx)
    build_md(md)
    print(f"OK: {xlsx}")
    print(f"OK: {md}")
